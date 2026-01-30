# -*- coding: utf-8 -*-

import os
import sys
import json
import time
import openpyxl
from openai import OpenAI

client = OpenAI(api_key=os.environ["OPENAI_API_KEY"])


# =========================
# API PROMPTS
# =========================

PROMPT_V = """あなたは、対象者と同じ職場で日常的に業務を見ている立場の、
落ち着いた現場の先輩です。

以下の月報全文を読み、
実際の月報フィードバック担当者が
要点を絞って、少し丁寧に書いたような
自然な月報草案を作成してください。

【制約】
- 敬体（です・ます）で書く
- 全体で5〜8行
- 肯定を中心にする
- 改善点に触れる場合は最大1点まで
- 原文に書かれていない事実や評価を追加しない
- 性格・能力への断定はしない
- 説教調にしない

【入力（月報全文）】
{TEXT}
"""

PROMPT_W = """以下の月報全文をもとに、
業務状況を第三者に共有するための
簡潔な要約報告を作成してください。

【制約】
- 2〜3行
- 事実ベースで記述する
- 評価・指導・感情表現は控える
- 特定の宛先（部署名・役職名など）は書かない
- 敬体（です・ます）で書く

【入力（月報全文）】
{TEXT}
"""


def call_api(prompt: str) -> str:
    res = client.responses.create(
        model=os.environ.get("OPENAI_MODEL", "gpt-5-mini"),
        input=prompt,
        max_output_tokens=int(os.environ.get("OPENAI_MAX_OUTPUT_TOKENS", "2000")),
        timeout=int(os.environ.get("OPENAI_TIMEOUT", "120")),
    )
    return res.output_text.strip()


# =========================
# STAGE 20
# =========================

def run_stage20(xlsx_path: str, out_path: str):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # 列定義（固定）
    COL_G = "G"  # 人間関係
    COL_I = "I"  # 積極性
    COL_K = "K"  # ヒヤリハット
    COL_M = "M"  # 本社相談

    COL_U = "U"  # 月報全文
    COL_V = "V"  # 月報草案
    COL_W = "W"  # 要約報告

    for row in range(2, ws.max_row + 1):
        g = ws[f"{COL_G}{row}"].value or ""
        i = ws[f"{COL_I}{row}"].value or ""
        k = ws[f"{COL_K}{row}"].value or ""
        m = ws[f"{COL_M}{row}"].value or ""

        full_text = "\n".join([g, i, k, m]).strip()
        ws[f"{COL_U}{row}"] = full_text

        if not full_text:
            continue

        # V列
        v_text = call_api(PROMPT_V.format(TEXT=full_text))
        ws[f"{COL_V}{row}"] = v_text

        # W列
        w_text = call_api(PROMPT_W.format(TEXT=full_text))
        ws[f"{COL_W}{row}"] = w_text

        time.sleep(1)

    wb.save(out_path)


if __name__ == "__main__":
    in_file = sys.argv[1]
    out_file = sys.argv[2]
    run_stage20(in_file, out_file)