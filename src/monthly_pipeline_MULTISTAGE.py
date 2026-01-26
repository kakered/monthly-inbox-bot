# -*- coding: utf-8 -*-
"""
monthly_pipeline_MULTISTAGE.py

Multi-stage monthly report pipeline.

Stages (intended):
  STAGE1: Build "prep" workbook (API前の整形). Creates blank feedback columns.
  STAGE2: Run API on prep workbook -> overview/per_person outputs (existing excel_exporter).
  (Future) STAGE3/4/5: accumulation & comparisons.

This file is designed to be *compatible* with the existing repository modules:
- src.dropbox_io.DropboxIO
- src.state_store.StateStore
- src.excel_exporter.process_monthly_workbook

Notes:
- This pipeline intentionally uses StateStore.save(dbx) / StateStore.load(dbx, path)
- DropboxIO must provide: list_folder(), download_to_bytes(), upload_bytes()/write_file_bytes()
"""

from __future__ import annotations

import os
import re
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.utils.datetime import from_excel

from .dropbox_io import DropboxIO
from .state_store import StateStore


# ---------------------------
# Config
# ---------------------------

@dataclass
class MonthlyMultiStageConfig:
    # Dropbox paths
    inbox_path: str = "/0-Inbox/monthlyreports"
    prep_out_dir: str = "/0-System/monthly_pipeline/prep"
    outbox_monthly_dir: str = "/0-Outbox/monthly"           # legacy (overview+per_person in one folder)
    outbox_overview_dir: str = "/0-Outbox/monthly_overview" # overview-only
    state_path: str = "/0-System/state.json"

    # Controls
    max_files_per_run: int = 1

    @classmethod
    def from_env(cls) -> "MonthlyMultiStageConfig":
        return cls(
            inbox_path=os.getenv("MONTHLY_INBOX_PATH", "/0-Inbox/monthlyreports"),
            prep_out_dir=os.getenv("MONTHLY_PREP_DIR", "/0-System/monthly_pipeline/prep"),
            outbox_monthly_dir=os.getenv("MONTHLY_OUTBOX_DIR", "/0-Outbox/monthly"),
            outbox_overview_dir=os.getenv("MONTHLY_OVERVIEW_DIR", "/0-Outbox/monthly_overview"),
            state_path=os.getenv("STATE_PATH", "/0-System/state.json"),
            max_files_per_run=int(os.getenv("MONTHLY_MAX_FILES", "1")),
        )


# ---------------------------
# Helpers: parsing input excel
# ---------------------------

# Input sheet expected: "月報"
INPUT_SHEET_NAME = "月報"

# Output sheet expected: "Sheet1" (matches provided xltx template)
OUTPUT_SHEET_NAME = "Sheet1"

# Mapping: input columns -> output "fulltext" columns.
# The *feedback* columns are created as empty strings.
# 最新仕様: 社員番号/一致確認用/就業先 は出力に残さない
OUTPUT_COLUMNS = [
    "Month",
    "Person_ID",
    "人間関係（報連相）_FullText",
    "人間関係（報連相）_Draft_Feedback",
    "仕事の量・質_FullText",
    "仕事の量・質_Draft_Feedback",
    "積極性_FullText",
    "積極性_Draft_Feedback",
    "責任性_FullText",
    "責任性_Draft_Feedback",
    "ヒヤリハットの抽出・分析・対策_FullText",
    "ヒヤリハットの抽出・分析・対策_Draft_Feedback",
    "就業先での業務改善提案_FullText",
    "就業先での業務改善提案_Draft_Feedback",
    "本社への相談・要望_FullText",
    "本社への相談・要望_Draft_Feedback",
]

INPUT_TO_OUTPUT_FULLTEXT = {
    "人間関係（報連相）": "人間関係（報連相）_FullText",
    "仕事の量・質": "仕事の量・質_FullText",
    "積極性": "積極性_FullText",
    "責任性": "責任性_FullText",
    "ヒヤリハットの抽出・分析・対策": "ヒヤリハットの抽出・分析・対策_FullText",
    "就業先での業務改善提案": "就業先での業務改善提案_FullText",
    "本社への相談・要望": "本社への相談・要望_FullText",
}

FEEDBACK_COLUMNS = [c for c in OUTPUT_COLUMNS if c.endswith("_Draft_Feedback")]


def _excel_month_to_str(v: Any) -> str:
    """
    Input '対象月' may be:
      - datetime/date
      - Excel serial (int/float)
      - string
    Output: 'YYYY-MM' (string)
    """
    if v is None:
        return "unknown-month"
    if isinstance(v, str):
        s = v.strip()
        # try normalize like '2025/09' or '2025-09'
        m = re.search(r"(\d{4})[/-](\d{1,2})", s)
        if m:
            return f"{m.group(1)}-{int(m.group(2)):02d}"
        return s or "unknown-month"
    # datetime/date
    try:
        import datetime as _dt
        if isinstance(v, (_dt.datetime, _dt.date)):
            y = v.year
            m = v.month
            return f"{y:04d}-{m:02d}"
    except Exception:
        pass
    # excel serial
    if isinstance(v, (int, float)):
        try:
            dt = from_excel(v)
            return f"{dt.year:04d}-{dt.month:02d}"
        except Exception:
            return str(v)
    return str(v)


def _read_input_rows(xlsx_bytes: bytes) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)
    if INPUT_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Input workbook has no sheet '{INPUT_SHEET_NAME}'. sheets={wb.sheetnames}")
    ws = wb[INPUT_SHEET_NAME]

    # header row = 1
    headers = []
    for c in range(1, ws.max_column + 1):
        headers.append(ws.cell(1, c).value)

    # build list of dict rows
    rows: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        row = {}
        empty = True
        for c, h in enumerate(headers, start=1):
            if h is None or h == "":
                continue
            val = ws.cell(r, c).value
            if val not in (None, ""):
                empty = False
            row[str(h)] = val
        if not empty:
            rows.append(row)
    return rows


def _make_prep_workbook(rows: List[Dict[str, Any]]) -> Tuple[bytes, str]:
    """
    Create 'prep' workbook for API.
    Returns (xlsx_bytes, month_str).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = OUTPUT_SHEET_NAME

    # header
    for col_idx, name in enumerate(OUTPUT_COLUMNS, start=1):
        ws.cell(1, col_idx).value = name

    month_str = "unknown-month"

    for i, r in enumerate(rows, start=2):
        # Month
        m = _excel_month_to_str(r.get("対象月"))
        month_str = month_str if month_str != "unknown-month" else m
        ws.cell(i, 1).value = m

        # Person_ID
        ws.cell(i, 2).value = r.get("ID")

        # FullText columns
        out_map: Dict[str, Any] = {}
        for in_col, out_col in INPUT_TO_OUTPUT_FULLTEXT.items():
            out_map[out_col] = r.get(in_col)

        # write fulltext + feedback blanks
        for col_idx, name in enumerate(OUTPUT_COLUMNS, start=1):
            if name in out_map:
                ws.cell(i, col_idx).value = out_map[name]
            elif name in FEEDBACK_COLUMNS:
                ws.cell(i, col_idx).value = ""  # blank for API fill
            # Month/Person_ID already written

    # freeze header
    ws.freeze_panes = "A2"

    out = BytesIO()
    wb.save(out)
    return out.getvalue(), month_str


# ---------------------------
# Stage1 / Stage2
# ---------------------------

def stage1_prep(dbx: DropboxIO, state: StateStore, cfg: MonthlyMultiStageConfig) -> None:
    items = dbx.list_folder(cfg.inbox_path)

    # filter xlsx only
    xlsx_items = [it for it in items if (getattr(it, "name", "") or "").lower().endswith(".xlsx")]
    if not xlsx_items:
        return

    # cap
    xlsx_items = xlsx_items[: cfg.max_files_per_run]

    for it in xlsx_items:
        src_path = getattr(it, "path", None) or getattr(it, "path_lower", None)
        if not src_path:
            continue

        key = (getattr(it, "path_lower", None) or src_path).lower()

        if state.is_processed(key):
            print(f"[MONTHLY][STAGE1] SKIP already processed: {src_path}")
            continue

        print(f"[MONTHLY][STAGE1] PREP: {src_path}")
        raw = dbx.download_to_bytes(src_path)
        rows = _read_input_rows(raw)
        prep_bytes, month_str = _make_prep_workbook(rows)

        ts = datetime.now().strftime("%Y%m%d-%H%M%S")
        prep_name = f"{getattr(it, 'name', 'monthly.xlsx')}__prep__{ts}.xlsx"
        prep_path = f"{cfg.prep_out_dir}/{prep_name}"

        dbx.write_file_bytes(prep_path, prep_bytes)
        print(f"[MONTHLY][STAGE1] Wrote: {prep_path}")

        state.mark_processed(key, rev=None)
        state.save(dbx)


def _iter_prep_files(dbx: DropboxIO, cfg: MonthlyMultiStageConfig) -> List[Any]:
    items = dbx.list_folder(cfg.prep_out_dir)
    return [it for it in items if getattr(it, "name", "").lower().endswith(".xlsx")]


def _infer_month_from_prep_bytes(prep_bytes: bytes) -> str:
    try:
        wb = openpyxl.load_workbook(BytesIO(prep_bytes), data_only=True)
        ws = wb[OUTPUT_SHEET_NAME] if OUTPUT_SHEET_NAME in wb.sheetnames else wb.active
        # A2 should be Month
        v = ws.cell(2, 1).value
        return _excel_month_to_str(v)
    except Exception:
        return "unknown-month"


def stage2_api(dbx: DropboxIO, state: StateStore, cfg: MonthlyMultiStageConfig) -> None:
    from .excel_exporter import process_monthly_workbook

    prep_files = _iter_prep_files(dbx, cfg)
    if not prep_files:
        return

    for it in prep_files[: cfg.max_files_per_run]:
        prep_path = getattr(it, "path", None) or getattr(it, "path_lower", None)
        if not prep_path:
            continue

        key = (getattr(it, "path_lower", None) or prep_path).lower()
        if state.is_processed(key):
            # already API'd
            continue

        prep_bytes = dbx.download_to_bytes(prep_path)
        month_str = _infer_month_from_prep_bytes(prep_bytes)

        print(f"[MONTHLY][STAGE2] OVERVIEW_API: {prep_path}")
        overview_bytes, per_person_bytes = process_monthly_workbook(
            workbook_bytes=prep_bytes,
            source_path=prep_path,
        )

        ts = datetime.now().strftime("%Y%m%d-%H%M%S")

        # overview-only folder (current)
        out_overview_path = f"{cfg.outbox_overview_dir}/{month_str}__overview__{ts}.xlsx"
        dbx.write_file_bytes(out_overview_path, overview_bytes)
        print(f"[MONTHLY][STAGE2] Wrote: {out_overview_path}")

        # optional legacy outputs if per_person_bytes is present
        if per_person_bytes:
            out_per_person_path = f"{cfg.outbox_monthly_dir}/{month_str}__per_person__{ts}.xlsx"
            dbx.write_file_bytes(out_per_person_path, per_person_bytes)
            print(f"[MONTHLY][STAGE2] Wrote: {out_per_person_path}")

        state.mark_processed(key, rev=None)
        state.save(dbx)


# ---------------------------
# Runner
# ---------------------------

def run_multistage(dbx: Optional[DropboxIO] = None, cfg: Optional[MonthlyMultiStageConfig] = None) -> None:
    """
    Default behavior:
      - STAGE1 then STAGE2 in same run (fast validation).
    Control:
      MONTHLY_STAGE=1 -> stage1 only
      MONTHLY_STAGE=2 -> stage2 only
      MONTHLY_STAGE=12 -> stage1 then stage2 (default)
    """
    dbx = dbx or DropboxIO.from_env()
    cfg = cfg or MonthlyMultiStageConfig.from_env()

    state = StateStore.load(dbx, cfg.state_path)

    stage = os.getenv("MONTHLY_STAGE", "12").strip()

    if stage == "1":
        stage1_prep(dbx, state, cfg)
        return
    if stage == "2":
        stage2_api(dbx, state, cfg)
        return

    # default: stage1 then stage2
    stage1_prep(dbx, state, cfg)
    stage2_api(dbx, state, cfg)
