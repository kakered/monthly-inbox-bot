# -*- coding: utf-8 -*-
"""
src/excel_exporter.py

Purpose
- Provide process_monthly_workbook() that monthly_pipeline_MULTISTAGE imports.
- Be tolerant to different calling conventions (bytes / in_path+out_path / args+kwargs).
- For now: "pass-through + light normalization" (no OpenAI calls here).
  It loads an .xlsx, ensures it is a valid workbook, and writes it back.

This file is intentionally defensive: it prevents pipeline crashes even if workbook is empty,
sheet names vary, or caller passes args in unexpected forms.
"""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Optional, Tuple, Union

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook


PathLike = Union[str, Path]


@dataclass
class ExportResult:
    ok: bool
    message: str
    rows_processed: int = 0
    sheets: int = 0


def _load_wb_from_bytes(data: bytes) -> Workbook:
    bio = BytesIO(data)
    wb = load_workbook(bio)
    return wb


def _save_wb_to_bytes(wb: Workbook) -> bytes:
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _coerce_paths_from_args_kwargs(args: Tuple[Any, ...], kwargs: Dict[str, Any]) -> Tuple[Optional[PathLike], Optional[PathLike]]:
    """
    Try to find input/output path candidates from arbitrary args/kwargs.
    """
    in_keys = ("in_path", "input_path", "src_path", "xlsx_path", "path", "local_path")
    out_keys = ("out_path", "output_path", "dst_path", "dest_path", "save_path")

    in_path = None
    out_path = None

    for k in in_keys:
        if k in kwargs and isinstance(kwargs[k], (str, Path)):
            in_path = kwargs[k]
            break

    for k in out_keys:
        if k in kwargs and isinstance(kwargs[k], (str, Path)):
            out_path = kwargs[k]
            break

    # If not in kwargs, infer from positional args
    # Common patterns:
    #   process_monthly_workbook(in_path, out_path, ...)
    #   process_monthly_workbook(in_path, ...)
    if in_path is None and len(args) >= 1 and isinstance(args[0], (str, Path)):
        in_path = args[0]
    if out_path is None and len(args) >= 2 and isinstance(args[1], (str, Path)):
        out_path = args[1]

    return in_path, out_path


def process_monthly_workbook(*args: Any, **kwargs: Any) -> Union[bytes, ExportResult, Dict[str, Any]]:
    """
    Main entrypoint imported by monthly_pipeline_MULTISTAGE.py.

    Supported call styles (tolerant):
    1) bytes -> bytes
       out_bytes = process_monthly_workbook(xlsx_bytes)

    2) file path(s) -> ExportResult
       res = process_monthly_workbook("in.xlsx", "out.xlsx")

    3) kwargs variants:
       res = process_monthly_workbook(in_path="in.xlsx", out_path="out.xlsx")

    If output path is not provided, it will overwrite the input path (safe default for pipeline temp files).
    """
    # Case A: first arg is bytes
    if len(args) >= 1 and isinstance(args[0], (bytes, bytearray)):
        data = bytes(args[0])
        wb = _load_wb_from_bytes(data)
        # minimal normalization hook (no-op for now)
        return _save_wb_to_bytes(wb)

    in_path, out_path = _coerce_paths_from_args_kwargs(args, kwargs)

    if in_path is None:
        # Nothing we can do; return a structured result rather than crashing.
        return ExportResult(ok=False, message="No input workbook provided (no bytes and no in_path).")

    in_p = Path(in_path)
    if out_path is None:
        out_p = in_p  # overwrite
    else:
        out_p = Path(out_path)

    if not in_p.exists():
        return ExportResult(ok=False, message=f"Input workbook not found on local FS: {in_p}")

    try:
        wb = load_workbook(in_p)
        sheets = len(wb.sheetnames)

        # minimal "touch": ensure workbook is writable by saving it.
        out_p.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_p)

        return ExportResult(ok=True, message="Workbook exported (pass-through).", rows_processed=0, sheets=sheets)

    except Exception as e:
        return ExportResult(ok=False, message=f"Failed to process workbook: {e!r}")